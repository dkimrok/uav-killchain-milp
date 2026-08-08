"""
killchain_auto.py
────────────────────────────────────────────────────────────────
Kill Chain 자동화 실행기
  - 단일 구역(legacy) 및 4구역(zone) 모드 지원
  - n_points_per_zone 파라미터 존재 시 4구역 모드 자동 활성화
  - killchain_input.xlsx 자동 생성
  - MILP 풀기 (killchain_solver.py 호출)
"""

import argparse, json, os, sys, random, math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
try:
    from killchain_solver import build_and_solve, print_result, draw_map, draw_timeline
except Exception:
    # 인스턴스 생성(generate_dataframes/save_excel)만 쓰는 경우 솔버/그리기 모듈 불필요
    build_and_solve = print_result = draw_map = draw_timeline = None

# ── 4구역 레이아웃 상수 ────────────────────────────────────────
# Zone1(0~50, 0~50) Zone2(50~100, 0~50)
# Zone3(0~50,50~100) Zone4(50~100,50~100)
ZONE_OFFSETS = {1: (0, 0), 2: (50, 0), 3: (0, 50), 4: (50, 50)}
ZONE_SIZE    = 50
ATK_DEPOTS   = {1: (0, 0), 2: (100, 0), 3: (0, 100), 4: (100, 100)}

# ─── 기본 시나리오 파라미터 ───────────────────────────────────
DEFAULT_SCENARIO = {
    "area_size":               100,
    # 4구역 모드 (n_points_per_zone 지정 시 활성화)
    "n_points_per_zone":       10,    # 구역당 정찰 지점 수
    "point_value_range":       [1, 10],
    # T_pp=1 고정: targets_per_point_range=[1,1]
    "targets_per_point_range": [1, 1],
    "target_value_range":      [1, 10],
    "target_window_range":     [3, 15],   # Δ_k (분)
    "target_offset_range":     [2, 8],    # 지점→표적 오프셋 (km)
    # 정찰 UAV (구역 무관, 전역 이동)
    "n_recon_uav":    3,
    "recon_start":    [0, 0],
    "recon_speed":    260/60,         # 4.333 km/min (260 km/h)
    "recon_max_dist": 260 * 4.6,     # 1196 km
    # 공격 UAV (4구역 고정, 구역별 1기)
    "attack_speed":    1110/60,       # 18.5 km/min (1110 km/h)
    "attack_max_dist": 1110 * 1.5,   # 1665 km
    "attack_weapons":  8,
    # 솔버
    "time_limit":  60,
    "random_seed": 42,
}


# ═══════════════════════════════════════════════════════════════
# 1. 파라미터 → 데이터프레임 생성
# ═══════════════════════════════════════════════════════════════
def generate_dataframes(cfg: dict):
    rng  = random.Random(cfg.get("random_seed", 42))
    area = cfg.get("area_size", 100)
    n_ppz = cfg.get("n_points_per_zone", None)  # 4구역 모드 트리거

    # ── 정찰 지점 ──────────────────────────────────────────────
    if "points" in cfg and cfg["points"]:
        df_points = pd.DataFrame(cfg["points"])

    elif n_ppz is not None:
        # 4구역 모드: zone 컬럼 포함
        vlo, vhi = cfg.get("point_value_range", [1, 10])
        rows = []; pid = 1
        for zone, (ox, oy) in ZONE_OFFSETS.items():
            margin = 0.1 * ZONE_SIZE
            for _ in range(int(n_ppz)):
                rows.append({
                    "point_id": f"P{pid}",
                    "zone":     zone,
                    "x":  round(rng.uniform(ox + margin, ox + ZONE_SIZE - margin), 1),
                    "y":  round(rng.uniform(oy + margin, oy + ZONE_SIZE - margin), 1),
                    "value": rng.randint(vlo, vhi),
                })
                pid += 1
        df_points = pd.DataFrame(rows)

    else:
        # legacy 단일 구역 모드
        n = cfg.get("n_points", 4)
        vlo, vhi = cfg.get("point_value_range", [1, 10])
        rows = []
        for i in range(1, n+1):
            rows.append({
                "point_id": f"P{i}",
                "x": round(rng.uniform(0.1*area, 0.9*area), 1),
                "y": round(rng.uniform(0.1*area, 0.9*area), 1),
                "value": rng.randint(vlo, vhi),
            })
        df_points = pd.DataFrame(rows)

    # ── 표적 ────────────────────────────────────────────────────
    if "targets" in cfg and cfg["targets"]:
        df_targets = pd.DataFrame(cfg["targets"])
    else:
        nt_lo, nt_hi = cfg.get("targets_per_point_range", [1, 1])
        tv_lo, tv_hi = cfg.get("target_value_range",      [1, 10])
        tw_lo, tw_hi = cfg.get("target_window_range",     [3, 15])
        off_lo, off_hi = cfg.get("target_offset_range",   [2, 8])
        rows = []; kid = 1

        for _, pr in df_points.iterrows():
            n_t = rng.randint(nt_lo, nt_hi)
            for _ in range(n_t):
                angle = rng.uniform(0, 2*math.pi)
                dist  = rng.uniform(off_lo, off_hi)
                if n_ppz is not None:
                    # 4구역: 표적이 같은 구역 내에 머물도록 클리핑
                    zone = int(pr.zone)
                    ox, oy = ZONE_OFFSETS[zone]
                    tx = max(ox + 1, min(ox + ZONE_SIZE - 1,
                                        pr.x + dist * math.cos(angle)))
                    ty = max(oy + 1, min(oy + ZONE_SIZE - 1,
                                        pr.y + dist * math.sin(angle)))
                else:
                    tx = max(0, min(area, pr.x + dist * math.cos(angle)))
                    ty = max(0, min(area, pr.y + dist * math.sin(angle)))
                rows.append({
                    "target_id":     f"T{kid}",
                    "x":             round(tx, 1),
                    "y":             round(ty, 1),
                    "value":         rng.randint(tv_lo, tv_hi),
                    "valid_minutes": rng.randint(tw_lo, tw_hi),
                    "point_id":      pr.point_id,
                })
                kid += 1
        df_targets = pd.DataFrame(rows)

    # ── 정찰 UAV (구역 무관) ───────────────────────────────────
    if "recon_uav" in cfg and cfg["recon_uav"]:
        df_recon = pd.DataFrame(cfg["recon_uav"])
    else:
        n = cfg.get("n_recon_uav", 3)
        sx, sy = cfg.get("recon_start", [0, 0])
        rows = []
        for i in range(1, n+1):
            rows.append({
                "uav_id":   f"R{i}",
                "start_x":  sx, "start_y": sy,
                "speed":    cfg.get("recon_speed",    260/60),
                "max_dist": cfg.get("recon_max_dist", 260*4.6),
            })
        df_recon = pd.DataFrame(rows)

    # ── 공격 UAV ───────────────────────────────────────────────
    if "attack_uav" in cfg and cfg["attack_uav"]:
        df_attack = pd.DataFrame(cfg["attack_uav"])
    elif n_ppz is not None:
        # 4구역 모드: zone별 1기, zone 컬럼 포함
        rows = []
        for zone, (sx, sy) in ATK_DEPOTS.items():
            rows.append({
                "uav_id":   f"A{zone}",
                "zone":     zone,
                "start_x":  sx, "start_y": sy,
                "speed":    cfg.get("attack_speed",    1110/60),
                "max_dist": cfg.get("attack_max_dist", 1110*1.5),
                "weapons":  cfg.get("attack_weapons",  8),
            })
        df_attack = pd.DataFrame(rows)
    else:
        # legacy
        n = cfg.get("n_attack_uav", 1)
        sx, sy = cfg.get("attack_start", [0, 0])
        rows = []
        for i in range(1, n+1):
            rows.append({
                "uav_id":   f"A{i}",
                "start_x":  sx, "start_y": sy,
                "speed":    cfg.get("attack_speed",    30),
                "max_dist": cfg.get("attack_max_dist", 200),
                "weapons":  cfg.get("attack_weapons",  3),
            })
        df_attack = pd.DataFrame(rows)

    return df_points, df_targets, df_recon, df_attack


# ═══════════════════════════════════════════════════════════════
# 2. 데이터프레임 → Excel 저장 (스타일 포함)
# ═══════════════════════════════════════════════════════════════
SHEET_FILL = {
    "points":     "D6E4F0",
    "targets":    "FCE4EC",
    "recon_uav":  "E8F5E9",
    "attack_uav": "FFF3E0",
}

def save_excel(df_points, df_targets, df_recon, df_attack, path: str):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_points.to_excel(writer,  sheet_name="points",     index=False)
        df_targets.to_excel(writer, sheet_name="targets",    index=False)
        df_recon.to_excel(writer,   sheet_name="recon_uav",  index=False)
        df_attack.to_excel(writer,  sheet_name="attack_uav", index=False)

    wb = load_workbook(path)
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    DAT_FONT = Font(name="Arial", size=10)
    CTR = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="AAAAAA")
    BRD = Border(left=thin, right=thin, top=thin, bottom=thin)

    for sname, df in [("points", df_points), ("targets", df_targets),
                      ("recon_uav", df_recon), ("attack_uav", df_attack)]:
        ws = wb[sname]
        row_fill = PatternFill("solid", fgColor=SHEET_FILL[sname])
        for col_idx in range(1, ws.max_column+1):
            c = ws.cell(row=1, column=col_idx)
            c.fill=HDR_FILL; c.font=HDR_FONT; c.alignment=CTR; c.border=BRD
        for row in ws.iter_rows(min_row=2):
            for c in row:
                c.font=DAT_FONT; c.alignment=CTR; c.border=BRD; c.fill=row_fill
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=6)
            ws.column_dimensions[get_column_letter(col[0].column)].width = mx+4
        ws.row_dimensions[1].height = 18
    wb.save(path)


# ═══════════════════════════════════════════════════════════════
# 3. 결과 요약 Excel
# ═══════════════════════════════════════════════════════════════
def save_result_excel(res, path: str):
    rec_rows = []
    for r in res['R']:
        for p, t in res['rec_routes'].get(r, []):
            rec_rows.append({"UAV": r, "Point": p,
                             "Arrival_min": round(t, 2),
                             "Value": res['p_val'][p]})
    df_rec = pd.DataFrame(rec_rows) if rec_rows else pd.DataFrame(
        columns=["UAV","Point","Arrival_min","Value"])

    hit_rows = []
    for k in res['K']:
        a   = res['hits'].get(k)
        pk  = res['k_point'][k]
        dp  = res['D_vals'][pk]
        dlt = res['k_win'][k]
        row = {"Target": k, "Recon_point": pk,
               "Value": res['k_val'][k],
               "Time_window_min": dlt,
               "Window_open":  round(dp, 2),
               "Window_close": round(dp+dlt, 2),
               "Hit":    "✓" if a else "✗",
               "Hit_by": a if a else "",
               "Hit_time": round(res['T_vals'].get((a,k), 0), 2) if a else ""}
        hit_rows.append(row)
    df_hit = pd.DataFrame(hit_rows)

    rec_val = sum(res['p_val'][p] for p in res['explored'])
    atk_val = sum(res['k_val'][k] for k in res['hits'])
    df_sum  = pd.DataFrame([{
        "Status":          res['status'],
        "Recon_value":     rec_val,
        "Strike_value":    atk_val,
        "Total_value":     rec_val + atk_val,
        "Points_explored": len(res['explored']),
        "Targets_hit":     len(res['hits']),
    }])

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_sum.to_excel(writer, sheet_name="summary",       index=False)
        df_rec.to_excel(writer, sheet_name="recon_result",  index=False)
        df_hit.to_excel(writer, sheet_name="strike_result", index=False)
    print(f"  결과 Excel 저장: {path}")


# ═══════════════════════════════════════════════════════════════
# 4. 메인 실행
# ═══════════════════════════════════════════════════════════════
def run(cfg: dict, out_dir: str = "."):
    os.makedirs(out_dir, exist_ok=True)
    print("\n" + "═"*60)
    print("  Kill Chain 자동화 실행기")
    print("═"*60)

    df_p, df_t, df_r, df_a = generate_dataframes(cfg)
    print(f"\n[입력 요약]")
    print(f"  정찰 지점 : {len(df_p)}개")
    print(f"  표적      : {len(df_t)}개")
    print(f"  정찰 UAV  : {len(df_r)}대")
    print(f"  공격 UAV  : {len(df_a)}대")

    xls_in = os.path.join(out_dir, "killchain_input.xlsx")
    save_excel(df_p, df_t, df_r, df_a, xls_in)
    print(f"\n  입력 Excel 생성: {xls_in}")

    print("\n[MILP 실행 중...]")
    res = build_and_solve(xls_in, time_limit=cfg.get("time_limit", 120), msg=False)
    print_result(res)

    draw_map(res,      save_path=os.path.join(out_dir, "killchain_map.png"))
    draw_timeline(res, save_path=os.path.join(out_dir, "killchain_timeline.png"))
    save_result_excel(res, os.path.join(out_dir, "killchain_result.xlsx"))
    print(f"\n[완료] 모든 파일이 '{out_dir}' 에 저장되었습니다.\n")
    return res


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Kill Chain MILP 자동화")
    parser.add_argument("--config", type=str, default=None)
    parser.add_argument("--out",    type=str, default="./kc_output")
    args = parser.parse_args()
    cfg = json.load(open(args.config)) if args.config else dict(DEFAULT_SCENARIO)
    run(cfg, out_dir=args.out)